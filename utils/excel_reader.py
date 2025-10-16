"""
Excel読み込みモジュール
素材マスター.xlsxと加工方法マスター.xlsxからデータを読み込む
"""
import openpyxl
from tkinter import filedialog, messagebox
from typing import Dict, List, Tuple, Optional
import os


class ExcelReader:
    """Excel読み込み処理を行うクラス"""
    
    def __init__(self):
        self.materials: Dict[str, str] = {}  # 表示名 -> 素材ID
        self.processing_methods: Dict[str, str] = {}  # 表示名 -> 加工ID
        self.materials_data: Dict[str, Dict[str, str]] = {}  # 素材IDの詳細データ
        self.processing_methods_data: Dict[str, Dict[str, str]] = {}  # 加工IDの詳細データ
        # 素材区分対応の新しい属性
        self.material_categories: Dict[str, List[str]] = {}  # 素材区分 -> 素材名リスト
        self.material_name_to_id: Dict[str, str] = {}  # 素材名 -> 素材ID
    
    def select_and_load_materials_file(self) -> bool:
        """素材マスターファイルを選択し、読み込む"""
        file_path = filedialog.askopenfilename(
            title="素材マスター.xlsxを選択してください",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=os.path.expanduser("~")
        )
        
        if not file_path:
            return False
        
        return self._load_materials_file(file_path)
    
    def select_and_load_processing_methods_file(self) -> bool:
        """加工方法マスターファイルを選択し、読み込む"""
        file_path = filedialog.askopenfilename(
            title="加工方法マスター.xlsxを選択してください",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=os.path.expanduser("~")
        )
        
        if not file_path:
            return False
        
        return self._load_processing_methods_file(file_path)
    
    def _load_materials_file(self, file_path: str) -> bool:
        """素材マスターファイルを読み込む"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            
            materials = {}  # display_name -> material_id のマッピング
            self.materials_data = {}  # material_id -> {name, description} の詳細データ
            
            # ヘッダー行（1行目）から列インデックスを取得
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            material_name_col = None
            material_id_col = None
            material_category_col = None
            
            # ヘッダー行から対象列を検索
            for col_idx, header_value in enumerate(header_row):
                if header_value:
                    header_str = str(header_value).strip()
                    if header_str == "素材名":
                        material_name_col = col_idx
                    elif header_str == "素材ID":
                        material_id_col = col_idx
                    elif header_str == "素材区分":
                        material_category_col = col_idx
            
            if material_name_col is None or material_id_col is None or material_category_col is None:
                messagebox.showerror("エラー", "素材マスターファイルに「素材名」「素材ID」または「素材区分」の列が見つかりません。")
                return False
            
            # 2行目からデータを読み込み
            self.material_categories = {}
            self.material_name_to_id = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > max(material_name_col, material_id_col, material_category_col):
                    material_name = row[material_name_col]
                    material_id = row[material_id_col]
                    material_category = row[material_category_col]
                    
                    if material_name and material_id and material_category:
                        material_name = str(material_name).strip()
                        material_id = str(material_id).strip()
                        material_category = str(material_category).strip()
                        
                        # 説明列があれば取得（素材名の隣の列を想定）
                        material_description = ""
                        if len(row) > material_name_col + 1 and row[material_name_col + 1]:
                            material_description = str(row[material_name_col + 1]).strip()
                        
                        if material_name and material_id and material_category:
                            # 素材区分ごとの素材名リストを構築
                            if material_category not in self.material_categories:
                                self.material_categories[material_category] = []
                            self.material_categories[material_category].append(material_name)
                            
                            # 素材名からIDへのマッピング
                            self.material_name_to_id[material_name] = material_id
                            
                            # ドロップダウン表示用の文字列を作成（互換性のため保持）
                            if material_description:
                                display_name = f"{material_name} - {material_description}"
                            else:
                                display_name = material_name
                            
                            materials[display_name] = material_id
                            self.materials_data[material_id] = {
                                'name': material_name,
                                'description': material_description,
                                'category': material_category
                            }
            
            if not materials:
                messagebox.showerror("エラー", "素材マスターファイルにデータが見つかりません。")
                return False
            
            self.materials = materials
            messagebox.showinfo("完了", f"素材マスターを読み込みました。({len(materials)}件)")
            return True
            
        except Exception as e:
            messagebox.showerror("エラー", f"素材マスターファイルの読み込みに失敗しました:\\n{str(e)}")
            return False
    
    def _load_processing_methods_file(self, file_path: str) -> bool:
        """加工方法マスターファイルを読み込む"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            
            processing_methods = {}  # display_name -> processing_id のマッピング
            self.processing_methods_data = {}  # processing_id -> {name, description} の詳細データ
            
            # ヘッダー行（1行目）から列インデックスを取得
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            method_name_col = None
            method_id_col = None
            
            # ヘッダー行から対象列を検索
            for col_idx, header_value in enumerate(header_row):
                if header_value:
                    header_str = str(header_value).strip()
                    if header_str == "加工方法名":
                        method_name_col = col_idx
                    elif header_str == "加工ID":
                        method_id_col = col_idx
            
            if method_name_col is None or method_id_col is None:
                messagebox.showerror("エラー", "加工方法マスターファイルに「加工方法名」または「加工ID」の列が見つかりません。")
                return False
            
            # 2行目からデータを読み込み
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > max(method_name_col, method_id_col):
                    method_name = row[method_name_col]
                    method_id = row[method_id_col]
                    
                    if method_name and method_id:
                        method_name = str(method_name).strip()
                        method_id = str(method_id).strip()
                        
                        # 説明列があれば取得（加工方法名の隣の列を想定）
                        method_description = ""
                        if len(row) > method_name_col + 1 and row[method_name_col + 1]:
                            method_description = str(row[method_name_col + 1]).strip()
                        
                        if method_name and method_id:
                            # ドロップダウン表示用の文字列を作成
                            if method_description:
                                display_name = f"{method_name} - {method_description}"
                            else:
                                display_name = method_name
                            
                            processing_methods[display_name] = method_id
                            self.processing_methods_data[method_id] = {
                                'name': method_name,
                                'description': method_description
                            }
            
            if not processing_methods:
                messagebox.showerror("エラー", "加工方法マスターファイルにデータが見つかりません。")
                return False
            
            self.processing_methods = processing_methods
            messagebox.showinfo("完了", f"加工方法マスターを読み込みました。({len(processing_methods)}件)")
            return True
            
        except Exception as e:
            messagebox.showerror("エラー", f"加工方法マスターファイルの読み込みに失敗しました:\\n{str(e)}")
            return False
    
    def get_materials_list(self) -> List[str]:
        """素材名のリストを取得"""
        return list(self.materials.keys())
    
    def get_processing_methods_list(self) -> List[str]:
        """加工方法名のリストを取得"""
        return list(self.processing_methods.keys())
    
    def get_material_code(self, display_name: str) -> Optional[str]:
        """表示名から素材IDを取得"""
        return self.materials.get(display_name)
    
    def get_processing_method_code(self, display_name: str) -> Optional[str]:
        """表示名から加工IDを取得"""
        return self.processing_methods.get(display_name)
    
    def get_material_details(self, material_id: str) -> Optional[Dict[str, str]]:
        """素材IDから詳細情報を取得"""
        return self.materials_data.get(material_id)
    
    def get_processing_method_details(self, method_id: str) -> Optional[Dict[str, str]]:
        """加工IDから詳細情報を取得"""
        return self.processing_methods_data.get(method_id)
    
    def is_ready(self) -> bool:
        """両方のマスターファイルが読み込まれているかチェック"""
        return bool(self.materials) and bool(self.processing_methods)
    
    def get_material_categories_list(self) -> List[str]:
        """素材区分のリストを取得"""
        return list(self.material_categories.keys())
    
    def get_materials_by_category(self, category: str) -> List[str]:
        """指定された素材区分に属する素材名のリストを取得"""
        return self.material_categories.get(category, [])
    
    def get_material_id_by_name(self, material_name: str) -> Optional[str]:
        """素材名から素材IDを取得"""
        return self.material_name_to_id.get(material_name)